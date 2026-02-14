#!/usr/bin/env python3
import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import openpyxl

SCHEMA_VERSION = "0.1.0"

def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")

def slug_id(*parts: str) -> str:
    s = "_".join([p.strip() for p in parts if p and str(p).strip()])
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "item"

def normalize_mark(cell: Any) -> str:
    if cell is None:
        return "none"
    s = str(cell).strip().lower()
    if s in ("x", "✗"):
        return "none"
    if s in ("✓", "check", "ok", "yes"):
        return "present"
    if "sidechain" in s:
        return "sidechain_in"
    if "links" in s or "left" in s:
        return "left"
    if "rechts" in s or "right" in s:
        return "right"
    if "✓" in s:
        if "links" in s:
            return "left"
        if "rechts" in s:
            return "right"
        return "present"
    return "unknown"

def write_json(path: Path, obj: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

def build_meta(*source_files: str) -> Dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at_utc": utc_now_iso(),
        "source_files": list(source_files),
    }

GEAR_COLUMNS = [
    "Hauptkategorie",
    "Unterkategorie",
    "Kategorie-Typ",
    "Hersteller",
    "Modell",
    "Anzahl",
    "Ein-/Ausgänge",
    "Parameter/Regler",
    "Strom/Info",
    "Notes/Highlights",
    "Besonderheiten / Technische Daten",
]

def convert_gear_xlsx(path: Path) -> Dict[str, Any]:
    df = pd.read_excel(path)
    missing = [c for c in GEAR_COLUMNS if c not in df.columns]
    if missing:
        raise SystemExit(f"[gear] Missing columns in {path.name}: {missing}")

    gear: List[Dict[str, Any]] = []
    for _, row in df.iterrows():
        manufacturer = str(row.get("Hersteller", "") or "").strip()
        model = str(row.get("Modell", "") or "").strip()

        item = {
            "id": slug_id(manufacturer, model),
            "main_category": str(row.get("Hauptkategorie", "") or "").strip(),
            "sub_category": str(row.get("Unterkategorie", "") or "").strip(),
            "category_type": str(row.get("Kategorie-Typ", "") or "").strip(),
            "manufacturer": manufacturer,
            "model": model,
            "count": int(row.get("Anzahl", 0) or 0),
            "io_text": str(row.get("Ein-/Ausgänge", "") or "").strip(),
            "controls_text": str(row.get("Parameter/Regler", "") or "").strip(),
            "power_text": str(row.get("Strom/Info", "") or "").strip(),
            "notes_text": str(row.get("Notes/Highlights", "") or "").strip(),
            "tech_text": str(row.get("Besonderheiten / Technische Daten", "") or "").strip(),
            "tags": [],
        }
        if item["manufacturer"] or item["model"] or item["main_category"]:
            gear.append(item)

    return {"meta": build_meta(path.name), "gear": gear}


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()

def _is_blank_row(ws, r: int, c1: int = 1, c2: int = 2) -> bool:
    # If first two relevant cells are empty => blank separator row for our matrices
    return _cell_str(ws.cell(r, c1).value) == "" and _cell_str(ws.cell(r, c2).value) == ""

def find_row_with_prefix(ws, col: int, prefixes: List[str]) -> Optional[int]:
    """Find first row where cell(row, col) starts with any prefix (case-insensitive)."""
    pref = [p.strip().lower() for p in prefixes]
    for r in range(1, ws.max_row + 1):
        s = _cell_str(ws.cell(r, col).value).lower()
        if not s:
            continue
        for p in pref:
            if s.startswith(p):
                return r
    return None

def parse_wide_matrix(ws, header_row: int, name_col: int = 1, chan_start_col: int = 2) -> Tuple[Dict[str, Any], int]:
    """
    Wide matrix layout (your current Patchbay Übersicht.xlsx):
      Row header_row:  [<matrix title in A>] [1] [2] [3] ...
      Next rows:       [<device name>]      [mark] [mark] ...
      Stops at first blank device name in col A (or blank separator row).

    Returns: (matrix_dict, last_row_used)
    """
    channels: List[int] = []
    c = chan_start_col
    while c <= ws.max_column:
        s = _cell_str(ws.cell(header_row, c).value)
        if s == "":
            break
        try:
            ch = int(float(s))
        except Exception:
            break
        channels.append(ch)
        c += 1

    if not channels:
        raise SystemExit(f"[patchbay] Wide matrix at row {header_row}: no channel numbers found (starting col {chan_start_col})")

    devices: List[Dict[str, Any]] = []
    r = header_row + 1
    while r <= ws.max_row:
        if _is_blank_row(ws, r, name_col, chan_start_col):
            break
        dev_name = _cell_str(ws.cell(r, name_col).value)
        if dev_name == "":
            break
        m: Dict[str, str] = {}
        for idx, ch in enumerate(channels):
            v = ws.cell(r, chan_start_col + idx).value
            m[str(ch)] = normalize_mark(v)
        devices.append({"name": dev_name, "map": m})
        r += 1

    if not devices:
        raise SystemExit(f"[patchbay] Wide matrix at row {header_row}: no device rows found under header")

    return {"channels": channels, "devices": devices}, (r - 1)


def find_matrix_header(ws) -> Optional[Tuple[int, int]]:
    target = "kanal"
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if s.startswith(target):
                return (r, c)
    return None

def read_device_headers(ws, header_row: int, start_col: int) -> List[Tuple[str, int]]:
    devices = []
    c = start_col + 1
    while c <= ws.max_column:
        name = ws.cell(header_row, c).value
        if name is None or str(name).strip() == "":
            break
        devices.append((str(name).strip(), c))
        c += 1
    return devices

def read_channels(ws, start_row: int, chan_col: int) -> List[int]:
    channels = []
    r = start_row
    while r <= ws.max_row:
        v = ws.cell(r, chan_col).value
        if v is None or str(v).strip() == "":
            break
        try:
            ch = int(v)
        except Exception:
            break
        channels.append(ch)
        r += 1
    return channels

def convert_patchbay_xlsx(path: Path) -> Dict[str, Any]:
    """
    Supports TWO layouts:

    (A) Wide matrix (your current file):
      Row 1: "Output Kanal Patchbay ..." + channel numbers across columns
      Row 8: "Input Kanal Patchbay ..."  + channel numbers across columns
      Devices are rows below each header; marks are in channel columns.

    (B) Legacy tall matrix:
      Row:   "Kanal" | <Device 1> | <Device 2> | ...
      Below: channel numbers down the Kanal column; marks are in device columns.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    def parse_matrix_from(header_row: int, chan_col: int) -> Tuple[Dict[str, Any], int]:
        devices = read_device_headers(ws, header_row, chan_col)
        if not devices:
            raise SystemExit(f"[patchbay] No device columns found on row {header_row} in {path.name}")

        first_chan_row = header_row + 1
        channels = read_channels(ws, first_chan_row, chan_col)
        if not channels:
            raise SystemExit(f"[patchbay] No channels found under 'Kanal' at row {header_row} in {path.name}")

        dev_objs = []
        for dev_name, dev_col in devices:
            m: Dict[str, str] = {}
            for idx, ch in enumerate(channels):
                r = first_chan_row + idx
                v = ws.cell(r, dev_col).value
                m[str(ch)] = normalize_mark(v)
            dev_objs.append({"name": dev_name, "map": m})

        matrix = {"channels": channels, "devices": dev_objs}
        last_row = first_chan_row + len(channels) - 1
        return matrix, last_row

    data: Dict[str, Any] = {"meta": build_meta(path.name)}

    # Prefer wide matrices if present (matches your current spreadsheet)
    out_row = find_row_with_prefix(ws, 1, ["output kanal patchbay"])
    in_row = find_row_with_prefix(ws, 1, ["input kanal patchbay"])

    if out_row:
        outputs, _ = parse_wide_matrix(ws, out_row, name_col=1, chan_start_col=2)
        data["outputs"] = outputs
    if in_row:
        inputs, _ = parse_wide_matrix(ws, in_row, name_col=1, chan_start_col=2)
        data["inputs"] = inputs

    # Fallback to legacy tall matrices if wide format not found
    if "outputs" not in data:
        # 1) Outputs (top)
        pos_out = find_matrix_header(ws)
        if not pos_out:
            raise SystemExit(f"[patchbay] Could not find wide matrix OR legacy header 'Kanal' in {path.name}")

        out_header_row, out_chan_col = pos_out
        outputs, out_last_row = parse_matrix_from(out_header_row, out_chan_col)
        data["outputs"] = outputs

        # 2) Inputs (bottom) — search for the next "Kanal" header BELOW outputs
        pos_in: Optional[Tuple[int, int]] = None
        for r in range(out_last_row + 1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                s = str(v).strip().lower()
                if s.startswith("kanal"):
                    pos_in = (r, c)
                    break
            if pos_in:
                break

        if pos_in:
            in_header_row, in_chan_col = pos_in
            inputs, _ = parse_matrix_from(in_header_row, in_chan_col)
            data["inputs"] = inputs
    return data

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--gear-xlsx", required=True, type=Path)
    ap.add_argument("--patchbay-xlsx", required=True, type=Path)
    ap.add_argument("--out-dir", required=True, type=Path)
    args = ap.parse_args()

    gear = convert_gear_xlsx(args.gear_xlsx)
    patch = convert_patchbay_xlsx(args.patchbay_xlsx)

    write_json(args.out_dir / "gear.json", gear)
    write_json(args.out_dir / "patchbay.json", patch)

    print("Wrote:", args.out_dir / "gear.json")
    print("Wrote:", args.out_dir / "patchbay.json")

if __name__ == "__main__":
    main()
